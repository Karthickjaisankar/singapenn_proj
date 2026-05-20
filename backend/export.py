from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf_report(state: dict) -> bytes:
    """Generate a PDF report with crime statistics and hotspots."""
    crimes = state.get("crimes", [])
    zones = state.get("zones", [])

    # Calculate stats
    total_crimes = len(crimes)
    severe_count = sum(1 for c in crimes if c.get("severity") == "severe")
    moderate_count = sum(1 for c in crimes if c.get("severity") == "moderate")
    low_count = sum(1 for c in crimes if c.get("severity") == "low")

    # Get top hotspots
    hotspots = {}
    for crime in crimes:
        place = crime.get("place_of_crime", "Unknown")
        hotspots[place] = hotspots.get(place, 0) + 1
    top_hotspots = sorted(hotspots.items(), key=lambda x: x[1], reverse=True)[:10]

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#1e40af"),
        spaceAfter=6,
        alignment=1,  # Center
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#1e40af"),
        spaceAfter=12,
        spaceBefore=12,
    )

    # Title
    story.append(Paragraph("SINGAPENE SCHEME", title_style))
    story.append(Paragraph("Crime Prevention & Police Patrol Optimization", styles["Normal"]))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.3 * inch))

    # Executive Summary
    story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
    summary_data = [
        ["Metric", "Value"],
        ["Total Crimes Analyzed", str(total_crimes)],
        ["Severe Crimes", f"{severe_count} ({severe_count*100//total_crimes if total_crimes else 0}%)"],
        ["Moderate Crimes", f"{moderate_count} ({moderate_count*100//total_crimes if total_crimes else 0}%)"],
        ["Low Risk Crimes", f"{low_count} ({low_count*100//total_crimes if total_crimes else 0}%)"],
        ["Patrol Zones", str(len(zones))],
        ["Data Period", "2022-2026"],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    # Top Hotspots
    story.append(PageBreak())
    story.append(Paragraph("TOP CRIME HOTSPOTS", heading_style))
    hotspot_data = [["Rank", "Location", "Crime Count"]]
    for i, (place, count) in enumerate(top_hotspots, 1):
        hotspot_data.append([str(i), place[:50], str(count)])
    hotspot_table = Table(hotspot_data, colWidths=[0.8 * inch, 3.5 * inch, 1.2 * inch])
    hotspot_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ]
        )
    )
    story.append(hotspot_table)
    story.append(Spacer(1, 0.3 * inch))

    # Patrol Zones
    story.append(Paragraph("PATROL ZONES", heading_style))
    zone_data = [["Zone ID", "Crime Count", "Severity Score"]]
    for zone in zones:
        zone_data.append([
            f"Zone {zone.get('zone_id', 'N/A')}",
            str(zone.get("crime_count", 0)),
            f"{zone.get('severity_score', 0):.1f}",
        ])
    zone_table = Table(zone_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    zone_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ]
        )
    )
    story.append(zone_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(state: dict) -> bytes:
    """Generate an Excel report with detailed crime data and statistics."""
    crimes = state.get("crimes", [])
    zones = state.get("zones", [])

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=12)

    ws_summary["A1"] = "SINGAPENE SCHEME - Crime Prevention Report"
    ws_summary["A1"].font = Font(bold=True, size=14, color="1e40af")
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws_summary["A4"] = "Key Statistics"
    ws_summary["A4"].font = Font(bold=True, size=12, color="1e40af")

    stats_data = [
        ("Total Crimes", len(crimes)),
        ("Severe Crimes", sum(1 for c in crimes if c.get("severity") == "severe")),
        ("Moderate Crimes", sum(1 for c in crimes if c.get("severity") == "moderate")),
        ("Low Risk Crimes", sum(1 for c in crimes if c.get("severity") == "low")),
        ("Patrol Zones", len(zones)),
        ("Data Period", "2022-2026"),
    ]

    row = 5
    for label, value in stats_data:
        ws_summary[f"A{row}"] = label
        ws_summary[f"B{row}"] = value
        ws_summary[f"A{row}"].font = Font(bold=True)
        row += 1

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # Sheet 2: Raw Crime Data
    ws_crimes = wb.create_sheet("Crimes")
    crime_headers = ["ID", "District", "Police Station", "Year", "FIR Number", "Crime Type", "Severity", "Place", "Date", "Hour", "Time Slot"]
    ws_crimes.append(crime_headers)

    for col_num, header in enumerate(crime_headers, 1):
        cell = ws_crimes.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for crime in crimes:
        ws_crimes.append([
            crime.get("id"),
            crime.get("district"),
            crime.get("police_station"),
            crime.get("year"),
            crime.get("fir_number"),
            crime.get("head"),
            crime.get("severity"),
            crime.get("place_of_crime"),
            crime.get("date_of_occurrence"),
            crime.get("hour"),
            crime.get("time_slot"),
        ])

    for col in ws_crimes.columns:
        ws_crimes.column_dimensions[col[0].column_letter].width = 15

    # Sheet 3: Top Hotspots
    ws_hotspots = wb.create_sheet("Top Hotspots")
    hotspots = {}
    for crime in crimes:
        place = crime.get("place_of_crime", "Unknown")
        hotspots[place] = hotspots.get(place, 0) + 1
    top_hotspots = sorted(hotspots.items(), key=lambda x: x[1], reverse=True)[:20]

    ws_hotspots.append(["Rank", "Location", "Crime Count"])
    for col_num, header in enumerate(["Rank", "Location", "Crime Count"], 1):
        cell = ws_hotspots.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for rank, (place, count) in enumerate(top_hotspots, 1):
        ws_hotspots.append([rank, place, count])

    ws_hotspots.column_dimensions["A"].width = 8
    ws_hotspots.column_dimensions["B"].width = 50
    ws_hotspots.column_dimensions["C"].width = 15

    # Sheet 4: Patrol Zones
    ws_zones = wb.create_sheet("Patrol Zones")
    ws_zones.append(["Zone ID", "Crime Count", "Severity Score", "Top Crime Spots"])
    for col_num, header in enumerate(["Zone ID", "Crime Count", "Severity Score", "Top Crime Spots"], 1):
        cell = ws_zones.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for zone in zones:
        top_spots = ", ".join([s.get("place", "")[:30] for s in zone.get("top_spots", [])[:3]])
        ws_zones.append([
            f"Zone {zone.get('zone_id')}",
            zone.get("crime_count"),
            f"{zone.get('severity_score', 0):.1f}",
            top_spots,
        ])

    ws_zones.column_dimensions["A"].width = 12
    ws_zones.column_dimensions["B"].width = 15
    ws_zones.column_dimensions["C"].width = 18
    ws_zones.column_dimensions["D"].width = 40

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
